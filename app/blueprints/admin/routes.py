from flask import render_template, request, flash, redirect, url_for, jsonify
from flask_login import login_required, current_user
from app.blueprints.admin import bp
from app.models.user import User, UserProfile
from app import db

def admin_required(f):
    """Decorator to require admin access"""
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Akses ditolak. Anda harus admin untuk mengakses halaman ini.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    """Admin dashboard"""
    return render_template('admin/dashboard.html')

@bp.route('/verifikasi')
@login_required
@admin_required
def verifikasi():
    """User verification interface"""
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    search_query = request.args.get('search', '')
    
    # Base query for users with profiles
    query = User.query.join(UserProfile).filter(UserProfile.id.isnot(None))
    
    # Apply status filter
    if status_filter == 'verified':
        query = query.filter(UserProfile.is_verified == True)
    elif status_filter == 'unverified':
        query = query.filter(UserProfile.is_verified == False)
    elif status_filter == 'complete':
        query = query.filter(UserProfile.verification_progress == 100)
    elif status_filter == 'incomplete':
        query = query.filter(UserProfile.verification_progress < 100)
    
    # Apply search filter
    if search_query:
        query = query.filter(
            db.or_(
                User.email.contains(search_query),
                UserProfile.nama_lengkap.contains(search_query),
                UserProfile.sekolah.contains(search_query)
            )
        )
    
    # Order by verification status and creation date
    users = query.order_by(
        UserProfile.is_verified.asc(),
        UserProfile.verification_progress.desc(),
        User.created_at.desc()
    ).all()
    
    # Calculate statistics
    total_users = User.query.join(UserProfile).count()
    verified_users = User.query.join(UserProfile).filter(UserProfile.is_verified == True).count()
    complete_profiles = User.query.join(UserProfile).filter(UserProfile.verification_progress == 100).count()
    pending_verification = User.query.join(UserProfile).filter(
        UserProfile.verification_progress == 100,
        UserProfile.is_verified == False
    ).count()
    
    stats = {
        'total_users': total_users,
        'verified_users': verified_users,
        'complete_profiles': complete_profiles,
        'pending_verification': pending_verification,
        'verification_rate': round((verified_users / total_users * 100) if total_users > 0 else 0, 1)
    }
    
    return render_template('admin/verifikasi.html', 
                         users=users, 
                         stats=stats,
                         status_filter=status_filter,
                         search_query=search_query)

@bp.route('/verifikasi/approve/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def approve_user(user_id):
    """Approve user verification"""
    user = User.query.get_or_404(user_id)
    
    if not user.profile:
        flash('User tidak memiliki profil.', 'danger')
        return redirect(url_for('admin.verifikasi'))
    
    if user.profile.verification_progress < 100:
        flash('Profil user belum lengkap. Tidak dapat diverifikasi.', 'warning')
        return redirect(url_for('admin.verifikasi'))
    
    user.profile.is_verified = True
    db.session.commit()
    
    flash(f'User {user.profile.get_display_name()} berhasil diverifikasi.', 'success')
    return redirect(url_for('admin.verifikasi'))

@bp.route('/verifikasi/reject/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def reject_user(user_id):
    """Reject user verification"""
    user = User.query.get_or_404(user_id)
    
    if not user.profile:
        flash('User tidak memiliki profil.', 'danger')
        return redirect(url_for('admin.verifikasi'))
    
    user.profile.is_verified = False
    db.session.commit()
    
    flash(f'Verifikasi user {user.profile.get_display_name()} ditolak.', 'warning')
    return redirect(url_for('admin.verifikasi'))

@bp.route('/verifikasi/bulk-action', methods=['POST'])
@login_required
@admin_required
def bulk_verification_action():
    """Handle bulk verification actions"""
    action = request.form.get('action')
    user_ids = request.form.getlist('user_ids')
    
    if not user_ids:
        flash('Tidak ada user yang dipilih.', 'warning')
        return redirect(url_for('admin.verifikasi'))
    
    success_count = 0
    error_count = 0
    
    for user_id in user_ids:
        try:
            user = User.query.get(int(user_id))
            if user and user.profile:
                if action == 'approve':
                    if user.profile.verification_progress == 100:
                        user.profile.is_verified = True
                        success_count += 1
                    else:
                        error_count += 1
                elif action == 'reject':
                    user.profile.is_verified = False
                    success_count += 1
        except (ValueError, AttributeError):
            error_count += 1
    
    db.session.commit()
    
    if success_count > 0:
        action_text = 'diverifikasi' if action == 'approve' else 'ditolak'
        flash(f'{success_count} user berhasil {action_text}.', 'success')
    
    if error_count > 0:
        flash(f'{error_count} user gagal diproses.', 'warning')
    
    return redirect(url_for('admin.verifikasi'))


@bp.route('/pembayaran')
@login_required
@admin_required
def pembayaran():
    """Payment approval interface"""
    from app.models.payment import Payment
    from app.models.registration import Registration
    from app.models.competition import Competition
    
    # Get filter parameters
    status_filter = request.args.get('status', 'pending')
    search_query = request.args.get('search', '')
    
    # Base query for payments with registrations
    query = Payment.query.join(Registration).join(Competition)
    
    # Apply status filter
    if status_filter != 'all':
        query = query.filter(Payment.status == status_filter)
    
    # Apply search filter
    if search_query:
        query = query.join(User).join(UserProfile).filter(
            db.or_(
                User.email.contains(search_query),
                UserProfile.nama_lengkap.contains(search_query),
                Competition.nama_kompetisi.contains(search_query)
            )
        )
    
    # Order by upload date (newest first)
    payments = query.order_by(Payment.tanggal_upload.desc()).all()
    
    # Calculate statistics
    total_payments = Payment.query.count()
    pending_payments = Payment.query.filter_by(status='pending').count()
    approved_payments = Payment.query.filter_by(status='approved').count()
    rejected_payments = Payment.query.filter_by(status='rejected').count()
    
    # Calculate total revenue
    total_revenue = db.session.query(db.func.sum(Payment.jumlah)).filter_by(status='approved').scalar() or 0
    
    stats = {
        'total_payments': total_payments,
        'pending_payments': pending_payments,
        'approved_payments': approved_payments,
        'rejected_payments': rejected_payments,
        'total_revenue': total_revenue,
        'approval_rate': round((approved_payments / total_payments * 100) if total_payments > 0 else 0, 1)
    }
    
    return render_template('admin/pembayaran.html', 
                         payments=payments, 
                         stats=stats,
                         status_filter=status_filter,
                         search_query=search_query)


@bp.route('/pembayaran/approve/<int:payment_id>', methods=['POST'])
@login_required
@admin_required
def approve_payment(payment_id):
    """Approve payment"""
    from app.models.payment import Payment
    
    payment = Payment.query.get_or_404(payment_id)
    notes = request.form.get('notes', '')
    
    if payment.status != 'pending':
        flash('Pembayaran ini sudah diproses.', 'warning')
        return redirect(url_for('admin.pembayaran'))
    
    success, message = payment.approve_payment(current_user.id, notes)
    
    if success:
        flash(f'Pembayaran untuk {payment.registration.competition.nama_kompetisi} berhasil disetujui.', 'success')
    else:
        flash(f'Gagal menyetujui pembayaran: {message}', 'error')
    
    return redirect(url_for('admin.pembayaran'))


@bp.route('/pembayaran/reject/<int:payment_id>', methods=['POST'])
@login_required
@admin_required
def reject_payment(payment_id):
    """Reject payment"""
    from app.models.payment import Payment
    
    payment = Payment.query.get_or_404(payment_id)
    notes = request.form.get('notes', '')
    
    if payment.status != 'pending':
        flash('Pembayaran ini sudah diproses.', 'warning')
        return redirect(url_for('admin.pembayaran'))
    
    if not notes:
        flash('Catatan penolakan harus diisi.', 'error')
        return redirect(url_for('admin.pembayaran'))
    
    success, message = payment.reject_payment(current_user.id, notes)
    
    if success:
        flash(f'Pembayaran untuk {payment.registration.competition.nama_kompetisi} ditolak.', 'warning')
    else:
        flash(f'Gagal menolak pembayaran: {message}', 'error')
    
    return redirect(url_for('admin.pembayaran'))


@bp.route('/pembayaran/bulk-action', methods=['POST'])
@login_required
@admin_required
def bulk_payment_action():
    """Handle bulk payment actions"""
    from app.models.payment import Payment
    
    action = request.form.get('action')
    payment_ids = request.form.getlist('payment_ids')
    notes = request.form.get('bulk_notes', '')
    
    if not payment_ids:
        flash('Tidak ada pembayaran yang dipilih.', 'warning')
        return redirect(url_for('admin.pembayaran'))
    
    if action == 'reject' and not notes:
        flash('Catatan penolakan harus diisi untuk aksi massal.', 'error')
        return redirect(url_for('admin.pembayaran'))
    
    success_count = 0
    error_count = 0
    
    for payment_id in payment_ids:
        try:
            payment = Payment.query.get(int(payment_id))
            if payment and payment.status == 'pending':
                if action == 'approve':
                    success, _ = payment.approve_payment(current_user.id, notes)
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                elif action == 'reject':
                    success, _ = payment.reject_payment(current_user.id, notes)
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
            else:
                error_count += 1
        except (ValueError, AttributeError):
            error_count += 1
    
    if success_count > 0:
        action_text = 'disetujui' if action == 'approve' else 'ditolak'
        flash(f'{success_count} pembayaran berhasil {action_text}.', 'success')
    
    if error_count > 0:
        flash(f'{error_count} pembayaran gagal diproses.', 'warning')
    
    return redirect(url_for('admin.pembayaran'))


@bp.route('/export')
@login_required
@admin_required
def export_dashboard():
    """Export dashboard"""
    from app.models.competition import Competition
    from app.models.registration import Registration
    from app.models.payment import Payment
    
    # Get competition statistics for export options
    competitions = Competition.query.all()
    
    export_stats = []
    for competition in competitions:
        registrations = Registration.query.filter_by(competition_id=competition.id).all()
        approved_registrations = [r for r in registrations if r.status == 'approved']
        paid_registrations = [r for r in registrations if r.payment and r.payment.status == 'approved']
        
        export_stats.append({
            'competition': competition,
            'total_registrations': len(registrations),
            'approved_registrations': len(approved_registrations),
            'paid_registrations': len(paid_registrations),
            'revenue': sum(r.payment.jumlah for r in registrations if r.payment and r.payment.status == 'approved')
        })
    
    return render_template('admin/export.html', export_stats=export_stats)


@bp.route('/export/participants/<int:competition_id>')
@login_required
@admin_required
def export_participants(competition_id):
    """Export participants for a specific competition"""
    from flask import make_response
    from app.models.competition import Competition
    from app.models.registration import Registration
    from app.models.payment import Payment
    
    competition = Competition.query.get_or_404(competition_id)
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')  # all, approved, paid
    format_type = request.args.get('format', 'excel')  # csv, excel
    
    # Build query based on filters
    query = Registration.query.filter_by(competition_id=competition_id)
    
    if status_filter == 'approved':
        query = query.filter_by(status='approved')
    elif status_filter == 'paid':
        query = query.join(Payment).filter(Payment.status == 'approved')
    
    registrations = query.all()
    
    # Prepare data
    headers = [
        'No', 'Nama Lengkap', 'Email', 'Sekolah', 'Kelas', 'NISN', 
        'WhatsApp', 'Instagram', 'Twitter', 'Tipe Registrasi', 
        'Status Registrasi', 'Tanggal Registrasi', 'Harga Terkunci',
        'Status Pembayaran', 'Tanggal Pembayaran', 'Nama Tim', 'Posisi Tim'
    ]
    
    data = []
    for i, registration in enumerate(registrations, 1):
        user = registration.user
        profile = user.profile if user.profile else None
        payment = registration.payment
        
        # Get team info if applicable
        team_name = ''
        team_position = ''
        if registration.team:
            team_name = registration.team.nama_tim
            # Find user's position in team
            for member in registration.team.members:
                if member.user_id == user.id:
                    team_position = member.get_position_display()
                    break
        
        row = [
            i,  # No
            profile.nama_lengkap if profile else '',
            user.email,
            profile.sekolah if profile else '',
            profile.kelas if profile else '',
            profile.nisn if profile else '',
            profile.whatsapp if profile else '',
            profile.instagram if profile else '',
            profile.twitter if profile else '',
            registration.get_type().title(),
            registration.get_status_display(),
            registration.tanggal_registrasi.strftime('%d/%m/%Y %H:%M'),
            f"Rp {registration.harga_terkunci:,}",
            payment.get_status_display() if payment else 'Belum ada pembayaran',
            payment.tanggal_upload.strftime('%d/%m/%Y %H:%M') if payment and payment.tanggal_upload else '',
            team_name,
            team_position
        ]
        data.append(row)
    
    # Generate filename
    status_suffix = f"_{status_filter}" if status_filter != 'all' else ""
    safe_competition_name = competition.nama_kompetisi.replace(' ', '_')
    
    if format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill
            import io
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Peserta {competition.nama_kompetisi}"
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"peserta_{safe_competition_name}{status_suffix}.xlsx"
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            format_type = 'csv'
    
    # CSV export (fallback or if explicitly requested)
    if format_type == 'csv':
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers and data
        writer.writerow(headers)
        writer.writerows(data)
        
        output.seek(0)
        filename = f"peserta_{safe_competition_name}{status_suffix}.csv"
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


@bp.route('/export/revenue')
@login_required
@admin_required
def export_revenue():
    """Export revenue report"""
    import csv
    import io
    from flask import make_response
    from app.models.competition import Competition
    from app.models.registration import Registration
    from app.models.payment import Payment
    
    # Get all competitions with revenue data
    competitions = Competition.query.all()
    
    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    headers = [
        'No', 'Nama Kompetisi', 'Kategori', 'Jenis', 'Harga Early Bird', 'Harga Reguler',
        'Total Registrasi', 'Registrasi Disetujui', 'Pembayaran Disetujui', 
        'Revenue Early Bird', 'Revenue Reguler', 'Total Revenue'
    ]
    writer.writerow(headers)
    
    # Write data
    total_revenue_all = 0
    for i, competition in enumerate(competitions, 1):
        registrations = Registration.query.filter_by(competition_id=competition.id).all()
        approved_registrations = [r for r in registrations if r.status == 'approved']
        paid_registrations = [r for r in registrations if r.payment and r.payment.status == 'approved']
        
        # Calculate revenue breakdown
        early_bird_revenue = 0
        regular_revenue = 0
        
        for registration in paid_registrations:
            if registration.harga_terkunci == competition.harga_early_bird:
                early_bird_revenue += registration.payment.jumlah
            else:
                regular_revenue += registration.payment.jumlah
        
        total_revenue = early_bird_revenue + regular_revenue
        total_revenue_all += total_revenue
        
        row = [
            i,
            competition.nama_kompetisi,
            competition.get_category_display(),
            competition.get_competition_type_display(),
            f"Rp {competition.harga_early_bird:,}",
            f"Rp {competition.harga_reguler:,}",
            len(registrations),
            len(approved_registrations),
            len(paid_registrations),
            f"Rp {early_bird_revenue:,}",
            f"Rp {regular_revenue:,}",
            f"Rp {total_revenue:,}"
        ]
        writer.writerow(row)
    
    # Add total row
    writer.writerow([])
    writer.writerow(['TOTAL', '', '', '', '', '', '', '', '', '', '', f"Rp {total_revenue_all:,}"])
    
    # Create response
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename="laporan_revenue_kompetisi.csv"'
    
    return response


@bp.route('/export/users')
@login_required
@admin_required
def export_users():
    """Export users report"""
    import csv
    import io
    from flask import make_response
    from app.models.user import User, UserProfile
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')  # all, verified, unverified
    
    # Build query based on filters
    query = User.query.join(UserProfile).filter(UserProfile.id.isnot(None))
    
    if status_filter == 'verified':
        query = query.filter(UserProfile.is_verified == True)
    elif status_filter == 'unverified':
        query = query.filter(UserProfile.is_verified == False)
    
    users = query.all()
    
    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    headers = [
        'No', 'Email', 'Nama Lengkap', 'Sekolah', 'Kelas', 'NISN', 
        'WhatsApp', 'Instagram', 'Twitter', 'Kelengkapan Profil (%)', 
        'Status Verifikasi', 'Tanggal Registrasi', 'Update Terakhir'
    ]
    writer.writerow(headers)
    
    # Write data
    for i, user in enumerate(users, 1):
        profile = user.profile
        
        row = [
            i,
            user.email,
            profile.nama_lengkap if profile else '',
            profile.sekolah if profile else '',
            profile.kelas if profile else '',
            profile.nisn if profile else '',
            profile.whatsapp if profile else '',
            profile.instagram if profile else '',
            profile.twitter if profile else '',
            profile.verification_progress if profile else 0,
            'Terverifikasi' if profile and profile.is_verified else 'Belum Terverifikasi',
            user.created_at.strftime('%d/%m/%Y %H:%M'),
            profile.updated_at.strftime('%d/%m/%Y %H:%M') if profile and profile.updated_at else ''
        ]
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    
    status_suffix = f"_{status_filter}" if status_filter != 'all' else ""
    filename = f"daftar_pengguna{status_suffix}.csv"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@bp.route('/kompetisi')
@login_required
@admin_required
def list_competitions():
    from app.models.competition import Competition
    competitions = Competition.query.order_by(Competition.tanggal_kompetisi.asc()).all()
    return render_template('admin/competition_list.html', competitions=competitions)

@bp.route('/kompetisi/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_competition(id):
    from app.models.competition import Competition
    from app.forms.competition import CompetitionEditForm
    competition = Competition.query.get_or_404(id)
    form = CompetitionEditForm(obj=competition)
    if form.validate_on_submit():
        form.populate_obj(competition)
        db.session.commit()
        flash('Kompetisi berhasil diperbarui.', 'success')
        return redirect(url_for('admin.list_competitions'))
    return render_template('admin/competition_edit.html', form=form, competition=competition)